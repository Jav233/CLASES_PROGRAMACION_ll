# Reimportar librerías necesarias después del reinicio del entorno
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Crear los datos
ids = [f"RF-{i:02d}" for i in range(1, 51)]
descripciones = [
    "Registrar un nuevo trabajo de reparación",
    "Capturar datos del cliente (nombre, correo, teléfono)",
    "Registrar tipo de dispositivo (PC, laptop, celular)",
    "Especificar marca y modelo del dispositivo",
    "Describir el problema reportado por el cliente",
    "Asignar número de orden único a cada trabajo",
    "Adjuntar imágenes del equipo ingresado",
    "Registrar fecha y hora de ingreso del equipo",
    "Registrar fecha estimada de entrega",
    "Generar comprobante de ingreso",
    "Asignar técnico responsable al trabajo",
    "Visualizar lista de trabajos asignados a cada técnico",
    "Registrar diagnóstico técnico del equipo",
    "Actualizar estado del trabajo (En proceso, Esperando repuesto, Finalizado)",
    "Registrar observaciones del técnico",
    "Asociar piezas o repuestos utilizados al trabajo",
    "Actualizar fecha de avance del trabajo",
    "Registrar prueba funcional después de la reparación",
    "Reasignar trabajo a otro técnico si es necesario",
    "Permitir al cliente ver el estado del trabajo en línea",
    "Permitir al cliente ver el historial de sus reparaciones",
    "Permitir al cliente calificar el servicio recibido",
    "Permitir al cliente enviar mensajes o comentarios",
    "Enviar notificaciones al cliente sobre el estado del trabajo",
    "Crear cuentas de usuario para empleados",
    "Establecer niveles de acceso por rol (Administrador, Técnico, Cliente)",
    "Permitir login seguro con usuario y contraseña",
    "Permitir recuperación de contraseña vía correo electrónico",
    "Generar reportes de trabajos por técnico",
    "Generar reportes de trabajos por fechas",
    "Generar reportes por tipo de dispositivo",
    "Visualizar estadísticas generales del sistema",
    "Exportar reportes a PDF o Excel",
    "Registrar insumos y repuestos en el inventario",
    "Descontar automáticamente insumos utilizados en una reparación",
    "Alertar cuando un insumo esté por agotarse",
    "Visualizar lista de insumos disponibles",
    "Generar presupuestos para los clientes",
    "Registrar pagos realizados por el cliente",
    "Generar comprobante de pago",
    "Registrar entregas de equipos reparados",
    "Registrar devoluciones de equipos sin reparación",
    "Generar historial de ingresos y egresos financieros",
    "Registrar actividad del sistema (bitácora de usuarios)",
    "Configurar parámetros del sistema (moneda, zona horaria, impuestos)",
    "Proteger el acceso mediante autenticación",
    "Guardar respaldo automático de la base de datos",
    "Cifrar la información confidencial",
    "Acceder a la plataforma desde dispositivos móviles",
    "Permitir escaneo de código QR para ver el estado del trabajo"
]
prioridades = [
    "Alta", "Alta", "Alta", "Alta", "Alta", "Alta", "Media", "Alta", "Media", "Media",
    "Alta", "Media", "Alta", "Alta", "Media", "Alta", "Media", "Alta", "Media", "Alta",
    "Media", "Media", "Media", "Alta", "Alta", "Alta", "Alta", "Alta", "Media", "Media",
    "Media", "Media", "Media", "Alta", "Alta", "Media", "Media", "Media", "Alta", "Media",
    "Alta", "Media", "Media", "Media", "Alta", "Alta", "Alta", "Alta", "Media", "Alta"
]
observaciones = [
    "Requerimiento clave para el flujo de trabajo",
    "Necesario para identificar y contactar al cliente",
    "Permite clasificar el tipo de equipo",
    "Ayuda a diferenciar dispositivos similares",
    "Importante para el diagnóstico inicial",
    "Evita duplicidad de trabajos",
    "Permite mejor seguimiento visual del equipo",
    "Registro cronológico del ingreso del equipo",
    "Da una idea clara del tiempo estimado al cliente",
    "Facilita la entrega del equipo",
    "Permite una asignación organizada de recursos",
    "Facilita la gestión y seguimiento de tareas",
    "Clave para iniciar diagnóstico y reparación",
    "Controla la trazabilidad del proceso de reparación",
    "Mejora la documentación del caso técnico",
    "Ayuda en la gestión del inventario",
    "Mantiene actualizado el historial de trabajo",
    "Garantiza la funcionalidad antes de la entrega",
    "Ofrece flexibilidad ante cambios de personal",
    "Mejora la experiencia del cliente",
    "Facilita la consulta de trabajos anteriores",
    "Proporciona retroalimentación del servicio",
    "Fomenta la comunicación directa",
    "Mantiene informado al cliente durante el proceso",
    "Permite acceso restringido a las funciones",
    "Controla acceso y operaciones dentro del sistema",
    "Brinda seguridad básica al sistema",
    "Mejora la usabilidad en caso de olvido de acceso",
    "Permite evaluar desempeño técnico",
    "Facilita análisis por temporadas",
    "Útil para determinar equipos con más fallas",
    "Brinda visión general de productividad",
    "Facilita entregas a entidades externas",
    "Controla el stock de insumos",
    "Automatiza el control de existencias",
    "Evita interrupciones por falta de materiales",
    "Da visibilidad rápida del stock disponible",
    "Ofrece una cotización clara y formal",
    "Registra la trazabilidad de ingresos",
    "Proporciona documento válido para el cliente",
    "Confirma la devolución exitosa del equipo",
    "Registra la cancelación del servicio",
    "Facilita auditorías y contabilidad",
    "Ayuda a controlar el uso del sistema",
    "Permite personalizar aspectos del sistema",
    "Incrementa la seguridad de acceso",
    "Protege la información ante fallos",
    "Protege la privacidad de los datos",
    "Mejora la accesibilidad desde cualquier lugar",
    "Agiliza la consulta de estado para el cliente"
]

# Crear DataFrame
df = pd.DataFrame({
    "ID": ids,
    "Descripción": descripciones,
    "Prioridad": prioridades,
    "Observaciones": observaciones
})

# Crear archivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Requerimientos Funcionales"

# Estilos
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True)
center_align = Alignment(horizontal="center", vertical="center")

alta_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
media_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
baja_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Agregar datos
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Aplicar estilos a encabezados
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = center_align

# Aplicar colores a prioridades
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        if cell.value == "Alta":
            cell.fill = alta_fill
        elif cell.value == "Media":
            cell.fill = media_fill
        elif cell.value == "Baja":
            cell.fill = baja_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align

# Ajustar anchos
for column_cells in ws.columns:
    max_length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

# Guardar archivo
path = "/mnt/data/Requerimientos_Funcionales_TecniLaptop_50.xlsx"
wb.save(path)

path
